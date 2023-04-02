import requests,json,time
from openpyxl import Workbook
from buff import get_buff_price
from update import update_price
from skinport import get_skinport_price
from currency_converter import CurrencyConverter
wb = Workbook()
ws = wb.active
c = CurrencyConverter()

with open("config.json","r",encoding="utf-8") as jconfig:
    config = json.load(jconfig)
    
cookies = {'session': config['settings']['cookies']}


def menu():
    while(True):
        print("\n" + "----------------------")
        print("PRESS [0] | Start script")
        print("PRESS [1] | Update prices (required at the first time)")
        print("PRESS [2] | Show Config")
        x = input("Enter your number: ")
        print( "----------------------" + "\n")
        if int(x) == 0:
            main()
        if int(x) == 1:
            update_price()
        if int(x) == 2:
            show_config()
        

def show_config():
    print("Cookie (cookie from buff)--> " + config['settings']['cookies'])
    print("Cookie (in Seconds) --> " + config['settings']['request_delay'])
    print("consolelog (True/False) --> " + config['settings']['consolelog'])

def main():
    t=0
    with open('weapons.txt',"r",encoding="utf-8") as f:
        for item in f:
            time.sleep(int(config['settings']['request_delay']))

            #get prices
            x0 = get_buff_price(item,cookies)
            x1 = get_skinport_price(item)

            #calculate profit and yuan in usd
            usd = c.convert(x0, 'CNY', 'USD')
            profit = (x1-usd)-(x1 * 0.12 )

            #prints console and saves in excel worksheet
            if usd != 0.0 and float(x1) != 0.0 and profit > 0:
                t += 1
                print(item.strip() + " ITEM " + str(t))
                
                if(bool(config['settings']['consolelog']) == True):
                    print("Buff: " + str(usd)+"$")
                    print("Skinport: " + str(x1) + "$")
                    print("Gewinn --> " + str(profit) + "$ -12%(inculdet)")
                    print("----------------------")
                ws[f'A{t+1}'] = str(item.strip())
                ws[f'B{t+1}'] = str(usd)
                ws[f'C{t+1}'] = str(x1)
                ws[f'D{t+1}'] = str(profit)
                wb.save("csgo_buff_skinport_pricing.xlsx")
menu()